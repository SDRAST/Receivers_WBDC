# -*- coding: utf-8 -*-
from openpyxl import load_workbook
from openpyxl.reader.excel import InvalidFileException
from openpyxl_support import *

import numpy
import logging

logger = logging.getLogger()
logger.setLevel(logging.DEBUG)

try:
  wb = load_workbook("43-2013-284.xlsx")
except IOError, details:
  logger.error(
      " loading spreadsheet failed with IO error.",
                        exc_info=True)
  raise RuntimeError
except InvalidFileException, details:
  logger.error(
      " .reader.excel doesn't like this file.",
                        exc_info=True)
  raise RuntimeError
except AttributeError, details:
  logger.error(
      " loading sheet failed with attribute error.",
                        exc_info=True)
  raise RuntimeError
sheet_names = wb.get_sheet_names()
logger.debug("sheet names: %s",
                      str(sheet_names))

data_ws = wb.get_sheet_by_name('Data')
column_names = get_column_names(data_ws)
logger.debug(" columns found:\n%s",
                      str(column_names))

highest_column_letter = get_column_letter(data_ws.get_highest_column())
logger.debug("highest column is %s", highest_column_letter)
highest_row_number = data_ws.get_highest_row()
logger.debug("Highest row is %d", highest_row_number)

data_range = "A2:"+highest_column_letter+str(highest_row_number)
logger.debug("Data range is %s", data_range)
data = numpy.array(data_ws.range(data_range))
